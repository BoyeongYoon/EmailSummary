"""
Author: Nancy Yoon (boyeong.nancy.yoon@gmail.com)
Date: June 16, 2022

Objectives: Convert the email list into an email summary text for each CustomerID and display the result
Constraints:
	1. The email summary text must conform to the following syntax:
		Primary_email_1[;Primary_email_2 to N][;Non_Primary_email_1 to N][;+ X more]
		where N and X are non-zero positive integer
	2. Primary emails must be displayed before the non primary emails
	3. The emails should be sorted in alphabetical order within each group (primary/non primary)
	4. The emails should be separated by a semi-colon (;)
	5. The email summary must be no more than 64 characters
	6. If the email summary text exceeds 64 characters, the emails should be removed from the text and be converted into [;+ X more]
	7. The email summary text should display as many email addresses as possible

How the program is implemented:
	(1) Open the given excel file and read data in it
	(2) Create a dictionary
			e.g. email_dict = {customer_id: [[primary email address], [non-primary email]]}
	(3) Sort each list alphabetically
	(4) Create an email summary
			e.g. Primary_email_1[;Primary_email_2 to N][;Non_Primary_email_1 to N][;+ X more]
	(5) Write an output in new excel file named output.xlsx
"""



# (1) Open the given excel file and read data in it
# openpyxl = a python library to read/write Excel
import openpyxl 
from openpyxl import Workbook
from openpyxl.styles import Font

path = "input.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active



# (2) Create a dictionary
# email_dict = {customer_id: [[primary email address], [non-primary email]]}
email_dict = {}

for i in range(2, sheet_obj.max_row):
	cell_1 = sheet_obj.cell(row = i, column = 1)
	cell_2 = sheet_obj.cell(row = i, column = 2)
	cell_3 = sheet_obj.cell(row = i, column = 3)

	customer_id = cell_1.value
	email = cell_2.value
	primary = cell_3.value


	if customer_id not in email_dict:
		email_dict[customer_id] = [[], []]

	if primary == 'yes':
			email_dict[customer_id][0].append(email)
	else:
			email_dict[customer_id][1].append(email)



# (3) Sort each list of email address (primary/non-primary) alphabetically
for email in email_dict.values():
  email[0].sort() # sort a list of primary email addresses
  email[1].sort() # sort a list of non-primary email addresses

for customer_id, email in email_dict.items():
  email_dict[customer_id] = email[0] + email[1]



# (4) Create an email summary
#	e.g. Primary_email_1[;Primary_email_2 to N][;Non_Primary_email_1 to N][;+ X more]e]
data = (("CustomerID", "EmailSummary"),)

for customer_id, email_list in email_dict.items():
	email_summary = email_list[0]
	for i in range(1, len(email_list)):
		if 64 - len(email_summary) >= len(email_list[i]):
			email_summary += ';' + email_list[i]
		else:
			email_summary += ';+ ' + str(len(email_list[i:])) + ' more'
			break
		
	data += (customer_id, email_summary),



# (5) Write an output in new excel file named output.xlsx
workbook = Workbook()
workbook.save(filename="email_summary.xlsx")

wb = openpyxl.load_workbook("email_summary.xlsx")
sheet = wb.active

for row in data: # e.g. row = alabama email1@alabama.com;email2@alabama.com; + 3 more
	sheet.append(row)

# Make bold font for the first row - CustomerID, EmailSummary
sheet['A1'].font = Font(bold=True)
sheet['B1'].font = Font(bold=True)
wb.save("output.xlsx")

